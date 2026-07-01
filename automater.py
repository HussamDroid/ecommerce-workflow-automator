import os
import time
import openpyxl
from pathlib import Path

# ==========================================
# CONFIGURATION
# ==========================================
EXCEL_FILE_PATH = r"C:\Path\To\Your\Master_Placeholder.xlsx"  # Update this to the exact path on the scanning PC
WATCH_DIRECTORY = r"Z:\Your\Network\Drive"
TARGET_STATUS = "Product already Given and Done"
POLL_INTERVAL = 60  # How often to check the folders (in seconds) -> 300s = 5 minutes

def process_folders():
    print(f"Scanning directory: {WATCH_DIRECTORY}...")
    
    # 1. Load the Excel workbook and the active sheet
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb.active
    except Exception as e:
        print(f"Failed to open Excel file. Is it currently open in another program? Error: {e}")
        return

    # 2. Dynamically find the column indexes for 'Barcode' and 'Notes (Sabbir)'
    barcode_col_idx = None
    status_col_idx = None

    # Assuming row 1 has the headers
    for cell in ws[1]:
        if cell.value == "Barcode":
            barcode_col_idx = cell.column
        elif cell.value == "Notes (Sabbir)":
            status_col_idx = cell.column

    if not barcode_col_idx or not status_col_idx:
        print("Error: Could not find 'Barcode' or 'Notes (Sabbir)' columns in the first row.")
        return

    # 3. Gather all folder names from the network drive
    # We do this first to minimize network calls and speed up the matching process
    existing_folders = set()
    for root, dirs, files in os.walk(WATCH_DIRECTORY):
        for directory in dirs:
            # Storing as string to ensure type matching with Excel data
            existing_folders.add(str(directory).strip())

    updates_made = False

    # 4. Iterate through the Excel rows and update status if it isn't already done
    # Starting from row 2 to skip headers
    for row in range(2, ws.max_row + 1):
        barcode_cell = ws.cell(row=row, column=barcode_col_idx)
        status_cell = ws.cell(row=row, column=status_col_idx)

        barcode_val = barcode_cell.value
        status_val = status_cell.value

        # Process if barcode exists AND the status is NOT already "Product already Given and Done"
        if barcode_val and status_val != TARGET_STATUS:
            # Convert Excel barcode to string (handling potential floats/ints)
            if isinstance(barcode_val, float):
                barcode_str = str(int(barcode_val)).strip()
            else:
                barcode_str = str(barcode_val).strip()
            
            # Check if this barcode exists as a folder name
            if barcode_str in existing_folders:
                ws.cell(row=row, column=status_col_idx).value = TARGET_STATUS
                updates_made = True
                
                # Handle blank/empty cells gracefully for the terminal printout
                old_status = status_val if status_val else "Empty/Blank"
                print(f"[UPDATED] Barcode {barcode_str}: '{old_status}' -> '{TARGET_STATUS}'")

    # 5. Save the file if updates were made
    if updates_made:
        try:
            wb.save(EXCEL_FILE_PATH)
            print(f"Success: Master.xlsx saved with new updates.")
        except PermissionError:
            print("Permission Error: Could not save the file. Please ensure it is closed and try again.")
    else:
        print("Scan complete. No new folders matched the criteria.")

def main():
    print(f"Starting Backend Workflow Automator...")
    print(f"Monitoring network drive: {WATCH_DIRECTORY} every {POLL_INTERVAL} seconds.")
    print("Press Ctrl+C to stop.")
    print("-" * 50)
    
    while True:
        try:
            process_folders()
        except KeyboardInterrupt:
            print("\nMonitoring stopped by user.")
            break
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
        
        # Wait for the next interval before scanning again
        time.sleep(POLL_INTERVAL)

if __name__ == "__main__":
    main()